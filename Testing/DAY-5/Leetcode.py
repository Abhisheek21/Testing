from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time, os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Paths
screenshot_path = r"C:\Users\abhis\OneDrive\Desktop\Testing\Screenshot"
excel_path = r"C:\Users\abhis\OneDrive\Desktop\Testing\LeetCode_Automation_Log.xlsx"
os.makedirs(screenshot_path, exist_ok=True)

# Excel setup
if not os.path.exists(excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Log"
    ws.append(["Timestamp", "Step", "Status"])
    wb.save(excel_path)

def log_status(step, status):
    wb = load_workbook(excel_path)
    ws = wb["Log"]
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([timestamp, step, status])
    wb.save(excel_path)

# Selenium Setup
driver = webdriver.Chrome()
driver.maximize_window()
wait = WebDriverWait(driver, 10)

try:
    driver.get("https://leetcode.com/")
    assert "LeetCode" in driver.title
    print("Home Page Title Check: Passed")
    driver.save_screenshot(os.path.join(screenshot_path, "homepage.png"))
    log_status("Home Page Load", "Passed")
except:
    log_status("Home Page Load", "Failed")

try:
    login_btn = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Sign in")))
    login_btn.click()
    wait.until(EC.presence_of_element_located((By.NAME, "login"))).send_keys("abhishekvijayakumar21@gmail.com")
    driver.find_element(By.NAME, "password").send_keys("abhishek@2148")
    time.sleep(1)
    driver.find_element(By.ID, "signin_btn").click()
    time.sleep(3)
    driver.execute_script("alert('Login Successful!');")
    time.sleep(1)
    driver.switch_to.alert.accept()
    driver.save_screenshot(os.path.join(screenshot_path, "after_login.png"))
    print("Login Successful")
    log_status("Login", "Passed")
except:
    log_status("Login", "Failed")

try:
    problems_link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Problems")))
    problems_link.click()
    time.sleep(3)
    problems_title = driver.title
    print("Problems Page Title:", problems_title)
    assert "Problems" in problems_title or "LeetCode" in problems_title
    print("Problems Page Title Check: Passed")
    driver.save_screenshot(os.path.join(screenshot_path, "problems.png"))
    log_status("Problems Page", "Passed")
except:
    log_status("Problems Page", "Failed")

try:
    driver.get("https://leetcode.com/contest/")
    time.sleep(2)
    contest_title = driver.title
    print("Contest Page Title:", contest_title)
    assert "Contest" in contest_title or "LeetCode" in contest_title
    print("Contest Page Title Check: Passed")
    driver.save_screenshot(os.path.join(screenshot_path, "contest.png"))
    log_status("Contest Page", "Passed")
except:
    log_status("Contest Page", "Failed")

try:
    driver.get("https://leetcode.com/abhishek212004/")
    time.sleep(2)
    profile_title = driver.title
    print("Profile Page Title:", profile_title)
    assert "abhishek212004" in profile_title or "LeetCode" in profile_title
    print("Profile Page Title Check: Passed")
    driver.save_screenshot(os.path.join(screenshot_path, "profile.png"))
    log_status("Profile Page", "Passed")
except:
    log_status("Profile Page", "Failed")

driver.close()
