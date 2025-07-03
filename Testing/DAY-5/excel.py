from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time
wb = load_workbook(r"C:\Users\abhis\OneDrive\Desktop\Testing\logins.xlsx")
sheet = wb.active
login_data = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
driver = webdriver.Chrome()  
driver.get('https://www.saucedemo.com/v1/')
driver.maximize_window()
time.sleep(2)
for username,password in login_data:
    driver.find_element(By.ID ,"user-name").send_keys(username)
    driver.find_element(By.ID,"password").send_keys(password)
    driver.find_element(By.ID,"login-button").click()
time.sleep(5)    
driver.quit()
