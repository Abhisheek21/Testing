from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import time

wb = load_workbook(r"C:\Users\abhis\OneDrive\Desktop\Python\cleartrip_flights (1).xlsx")
sheet = wb.active

driver = webdriver.Firefox()  
driver.get("https://www.cleartrip.com/")
driver.maximize_window()
time.sleep(3)

try:
    driver.find_element(By.ID, "wzrk-cancel").click()
except:
    pass

for row in sheet.iter_rows(min_row=2, values_only=True):
    from_city, to_city, flight_date = row

    driver.get("https://www.cleartrip.com/flights")
    time.sleep(4)

    from_input = driver.find_element(By.XPATH, "//input[@placeholder='Where from?']")
    from_input.clear()
    from_input.send_keys(from_city)
    time.sleep(1)
    from_input.send_keys(Keys.ENTER)
    time.sleep(1)
    
    to_input = driver.find_element(By.XPATH, "//input[@placeholder='Where to?']")
    to_input.clear()
    to_input.send_keys(to_city)
    time.sleep(1)
    to_input.send_keys(Keys.ENTER)
    time.sleep(1)

    date_input = driver.find_element(By.XPATH, "//div[@class='flex flex-middle p-relative homeCalender']/button")
    date_input.click()
    time.sleep(2)

    driver.execute_script(f"""
        document.querySelector('input[placeholder="Depart"]').value = "{flight_date}";
    """)

    search_button = driver.find_element(By.XPATH, "//button[@class='flex flex-middle btn-primary btn']")
    search_button.click()

    time.sleep(10)

    print(f"\nFlight availability for {from_city} to {to_city} on {flight_date}")
    try:
        flights = driver.find_elements(By.XPATH, "//div[contains(@class,'airlineInfo')]/span")[:3]
        for f in flights:
            print(f.text)
    except:
        print("No flights found or data not loaded properly.")

    print("-" * 40)
    time.sleep(5)

driver.quit()
